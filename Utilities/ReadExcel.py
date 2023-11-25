import openpyxl


def getData():
    workbook = openpyxl.load_workbook(r"C:\Users\MY PC\Desktop\login.xlsx")
    sheet = workbook["LoginData"]
    total_row = sheet.max_row
    total_col = sheet.max_column

    getdata = []
    for r in range(2, total_row + 1):
        list_Data = []
        for col in range(1, total_col + 1):
            data = sheet.cell(row=r, column=col).value
            list_Data.append(data)
        getdata.append(tuple(list_Data))

    return getdata
